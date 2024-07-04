import os
import ast
import sys
import logging
import math
import time
import pickle
import numpy as np
import pandas as pd
import scipy.stats as st
from statsmodels.stats.proportion import proportion_confint

from openpyxl import Workbook, load_workbook
from sklearn.metrics import accuracy_score

from phenoscore.phenoscorer import PhenoScorer
from phenoscore.explainability_lime.LIME import \
         LIMEConfiguration, OptiLIMEConfiguration


class GenerateAndAddResultsToFilePerSyndrome:
    """This class is used to generate and add results to an Excel file per syndrome."""
    def __init__(self, syndrome: str, lime_type: str, data_path: str, file_name_info: str,
                 lime_config: LIMEConfiguration, optilime_config: OptiLIMEConfiguration,
                 calculate_csi_vsi: bool, use_precomputed_kw: bool = False):
        """
        

        Parameters:
        - syndrome (str): The syndrome
        - lime_type (str): The type of LIME used for explanations: LIME or OptiLIME.
        - data_path (str): Path to the original data with two mandatory columns: 'hpo_all' and 'y_label'.
        - file_name_info (str): Info to include in file names, like the distance function (BMA-Resnik).
        - lime_config (LIMEConfiguration): Configuration for LIME and OptiLIME.
        - optilime_config (OptiLIMEConfiguration): Configuration for OptiLIME.
        - calculate_csi_vsi (bool): Whether to calculate the CSI and VSI or the other metrics
          (Weighted-R^2, explanation fidelity, and -accuracy).
        - use_precomputed_kw (bool): Whether to use a precomputed kernel width that is in the existing excel file.
        """
        self.syndrome = syndrome
        self.lime_type = lime_type
        self.data_path = data_path
        self.file_name_info = file_name_info
        self.full_path = rf'{self.syndrome}_{self.lime_type}\{self.syndrome}_{self.lime_type}_{self.file_name_info}.xlsx'
        print(f'The results are saved in an Excel file at the following location: {self.full_path}')
        self.lime_config = lime_config
        self.optilime_config = optilime_config
        self.calculate_csi_vsi = calculate_csi_vsi
        self.use_precomputed_kw = use_precomputed_kw

        # Setup logger
        logger = logging.getLogger('PhenoScore')
        logger.setLevel(logging.DEBUG)
        handler = logging.FileHandler('PhenoScore.txt')
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        
    def generate_results(self) -> None:
        """Calculate metrics for each patient and add results to the Excel file."""
        mode = 'hpo'
        phenoscorer = PhenoScorer(gene_name=self.syndrome, mode=mode,
                                  method_hpo_similarity='Resnik',
                                  method_summ_hpo_similarities='BMA')
        x, y, _, _ = phenoscorer.load_and_process_data_from_excel(self.data_path)
        x_copy, y_copy = np.copy(x), np.copy(y)
        for i, _ in enumerate(x):
            label, patient = y[i], x[i][0]
            logger = logging.getLogger('PhenoScore')
            logger.info("Processing patient %s for %s, label %s, %s with %s...",
                        i, self.syndrome, label, self.lime_type, self.file_name_info)
            logger.info("LIME config: %s, OptiLIME config: %s", self.lime_config, self.optilime_config)
            temp_data_x, temp_data_y = np.copy(x_copy), np.copy(y_copy)
            if label == 1:
                temp_data_x = np.delete(temp_data_x, [i, i+1], axis=0)
                temp_data_y = np.delete(temp_data_y, [i, i+1], axis=0)
            else:
                temp_data_x = np.delete(temp_data_x, [i, i-1], axis=0)
                temp_data_y = np.delete(temp_data_y, [i, i-1], axis=0)
            kernel_width = self.get_kernel_width(i, 2) if self.use_precomputed_kw else None
            if kernel_width:
                print(f"Using precomputed kernel width: {kernel_width}")
            results = phenoscorer.predict_new_sample(temp_data_x, temp_data_y, None, patient,
                                                     self.lime_config, self.optilime_config,
                                                     lime_iter=1, kernel_width=kernel_width)
            if self.calculate_csi_vsi:
                data = [i, phenoscorer.vus_results['vsi'], phenoscorer.vus_results['csi']]
                self.update_vsi_csi_in_excel(data)
            else:
                filename = rf'{self.syndrome}_{self.lime_type}\index_{i}_{self.lime_type}_{self.file_name_info}.pdf'
                results.gen_vus_figure(filename=filename)
                data = [i, label, phenoscorer.vus_results['kernel_width'],
                        phenoscorer.vus_results['max_rsquared'], phenoscorer.vus_results['sum_coeffs'],
                        0, 0, int(phenoscorer.vus_results['sum_coeffs'] > 0),
                        float(phenoscorer.vus_results['preds_hpo'][0])]
                self.add_row_to_excel(data, phenoscorer.vus_results['exp_hpos_all'])
            del temp_data_x, temp_data_y
            os.remove('yss.npy')
            os.remove('scaled_data.npy')
            os.remove('data_row.npy')
            os.remove('distances.npy')


    def get_kernel_width(self, row_index: int, column_index: int) -> float:
        " ""Get the kernel width from the Excel file at the given row and column index."
        df = pd.read_excel(self.full_path)
        kernel_width = df.iloc[row_index, column_index]
        return float(kernel_width)


    def add_row_to_excel(self, data: list, explanation_object) -> None:
        "Add a row to the Excel file with the given data and save the LIME explanation object separately."
        excel_path = self.full_path
        directory_path = rf'{self.syndrome}_{self.lime_type}'
        pickle_path = rf'{self.syndrome}_{self.lime_type}\{self.syndrome}_{self.lime_type}_{self.file_name_info}_index_{data[0]}.pkl'

        if not os.path.exists(directory_path):
            os.makedirs(directory_path)
        try:
            try:
                wb = load_workbook(excel_path)
            except FileNotFoundError:
                wb = Workbook()
                wb.active.append(['Index', 'Y Label', 'kernel width', 'weighted R-squared','sum_coeffs','CSI','VSI',f'classification_{self.lime_type}','classification_PhenoScore'])
            ws = wb.active
            with open(pickle_path, 'wb') as f:
                pickle.dump(explanation_object, f)
            ws.append(data)
            wb.save(excel_path)
            print("Row added successfully!")
        except Exception as e:
            print("An error occurred:", e)


    def update_vsi_csi_in_excel(self, data: list) -> None:
        " Add the calculated VSI and CSI in the Excel file."
        excel_path = self.full_path
        directory_path = rf'{self.syndrome}_{self.lime_type}'
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == data[0]: 
                    ws.cell(row=i, column=6, value=data[2]) #CSI
                    ws.cell(row=i, column=7, value=data[1]) #VSI
                    break
            wb.save(excel_path)
            print("VSI and CSI updated successfully!")
        except Exception as e:
            print("An error occurred:", e)

    def calculate_ci_and_average_metrics(self, **kwargs) -> None:
        """
        Calculate the confidence interval and average for specified metrics (for now the weighted R-squared, CSI, and the VSI).
        Accepts metrics as keyword arguments where keys are the metric names
        and values are the column names in the DataFrame.
        """
        df = pd.read_excel(self.full_path)

        for metric_name, column_name in kwargs.items():
            metric_values = df[column_name]
            metric_mean = metric_values.mean()
            metric_ci = st.t.interval(0.95, len(metric_values)-1, loc=metric_mean, scale=st.sem(metric_values))
            print(f"Confidence interval and mean of the {metric_name} for {self.syndrome}: {round(metric_mean, 2)} [{(round(metric_ci[0], 2), round(metric_ci[1], 2))}]")



    def calculate_ci_and_explanation_fidelity_accuracy(self,lime_type) -> None:
        " Calculate the confidence interval and average for explanation fidelity and -accuracy."
        df = pd.read_excel(self.full_path)
        df['classification_PhenoScore'] = np.where(df['classification_PhenoScore'] >= 0.5, 1, 0)
        df[f'classification_{lime_type}'] = np.where(df['sum_coeffs'] >= 0, 1, 0)
        self._calculate_and_print_metric(df, 'classification_PhenoScore', 'fidelity')
        self._calculate_and_print_metric(df, 'Y Label', 'accuracy')

    def _calculate_and_print_metric(self, df, column_name, metric_name) -> None:
        "helper function for calculate_CI_and_explanation_fidelity_accuracy()"
        acc = accuracy_score(df[column_name], df[f'classification_{self.lime_type}'])
        lower, upper = proportion_confint(acc*len(df), len(df), 0.05) # correct preds, number of instances, 95% CI.
        print(f"Explanation {metric_name} for {self.syndrome}: {acc:.2f} with 95% CI [{lower, upper}]")        
